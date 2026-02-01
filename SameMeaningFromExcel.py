import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

# ---------------- SELECT EXCEL FILE ----------------
root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(
    title="Select Excel File",
    filetypes=[("Excel files", "*.xlsx;*.xlsm")]
)

root.destroy()

if not file_path:
    raise SystemExit("No file selected")

# ---------------- READ EXCEL ----------------
df = pd.read_excel(file_path)

wb = load_workbook(file_path)
ws = wb.active

# ---------------- SAME MEANING DICTIONARY ----------------
payment_dict = {
    "cash": 1,
    "rakam": 1,
    "paisa": 1
}

# ---------------- CREATE PaymentID COLUMN IF NOT EXISTS ----------------
headers = [cell.value for cell in ws[1]]

if "PaymentID" in headers:
    id_col = headers.index("PaymentID") + 1
else:
    id_col = ws.max_column + 1
    ws.cell(row=1, column=id_col).value = "PaymentID"

# ---------------- SCAN ALL CELLS ----------------
for col_index, column in enumerate(df.columns, start=1):
    for row_index, value in enumerate(df[column], start=2):

        if pd.isna(value):
            continue

        text = str(value).strip().lower()

        if text in payment_dict:
            ws.cell(row=row_index, column=id_col).value = payment_dict[text]

# ---------------- SAVE SAME FILE ----------------
wb.save(file_path)

# ---------------- OPEN FILE ----------------
os.startfile(file_path)