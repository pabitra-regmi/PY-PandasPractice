import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import os #(added)


# Open file dialog
root = Tk()
root.withdraw()


file_path = filedialog.askopenfilename(
    title="Select Excel File",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

# Red fill for highlighting
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# Date regex patterns
gregorian_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")   # yyyy-mm-dd
bs_pattern = re.compile(r"^\d{4}\.\d{2}\.\d{2}$")        # yyyy.mm.dd

def is_valid_date(value):
    """Check if value is in yyyy-mm-dd OR yyyy.mm.dd format."""
    if value is None:
        return False
    value_str = str(value).strip()
    return bool(gregorian_pattern.fullmatch(value_str) or bs_pattern.fullmatch(value_str))

if file_path:
    try:
        wb = load_workbook(file_path)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Read headers
            headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
            col_map = {header: idx + 1 for idx, header in enumerate(headers)}
            
            print(f"Processing sheet: {sheet_name}")

            # ==========================
            # FIRST PASS — CLEAR COLORS
            # ==========================
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)

            # ==========================
            # SECOND PASS — VALIDATION
            # ==========================

            # Track duplicates
            unique_loan_acc = set()
            unique_accno = set()

            #(added)
            unique_membername = set()
            unique_memberid = set()

            for row in range(2, ws.max_row + 1): #

                # InterestRate validation
                if "interestrate" in col_map:
                    interest = ws.cell(row=row, column=col_map["interestrate"]).value
                    if interest is None or str(interest).strip() == "" or not str(interest).replace('.', '', 1).isdigit():
                        ws.cell(row=row, column=col_map["interestrate"]).fill = red_fill

                # Category validation
                if "category" in col_map:
                    cat_val = ws.cell(row=row, column=col_map["category"]).value
                    if cat_val is None or str(cat_val).strip() == "":
                        ws.cell(row=row, column=col_map["category"]).fill = red_fill

                # DepositTypeCode not null
                if "deposittypecode" in col_map:
                    dtcode = ws.cell(row=row, column=col_map["deposittypecode"]).value
                    if dtcode is None or str(dtcode).strip() == "":
                        ws.cell(row=row, column=col_map["deposittypecode"]).fill = red_fill

                # Duration validation
                if "duration" in col_map and "category" in col_map:
                    duration = ws.cell(row=row, column=col_map["duration"]).value
                    category = ws.cell(row=row, column=col_map["category"]).value
                    cat_str = str(category).strip().lower() if category else ""
                    if cat_str != "normal savings":
                        if duration is None or not str(duration).isdigit():
                            ws.cell(row=row, column=col_map["duration"]).fill = red_fill

                # DurationType validation
                if "durationtype" in col_map:
                    dtype = ws.cell(row=row, column=col_map["durationtype"]).value
                    dtype_str = str(dtype).strip().upper() if dtype else ""
                    if dtype_str not in ["Y", "M", "D"]:
                        ws.cell(row=row, column=col_map["durationtype"]).fill = red_fill

                # PeriodType validation
                if "periodtype" in col_map:
                    ptype = ws.cell(row=row, column=col_map["periodtype"]).value
                    ptype_str = str(ptype).strip().upper() if ptype else ""
                    if ptype_str not in ["Y", "M", "D"]:
                        ws.cell(row=row, column=col_map["periodtype"]).fill = red_fill

                # Period validation
                if "period" in col_map:
                    period = ws.cell(row=row, column=col_map["period"]).value
                    if period is None or str(period).strip() == "" or not str(period).isdigit():
                        ws.cell(row=row, column=col_map["period"]).fill = red_fill

                # Date column validation
                date_cols = ["accountopenonbs", "maturityonbs", "loanissuedate bs", "maturitydatebs"]
                for dc in date_cols:
                    if dc in col_map:
                        date_value = ws.cell(row=row, column=col_map[dc]).value
                        if not is_valid_date(date_value):
                            ws.cell(row=row, column=col_map[dc]).fill = red_fill

                # ================================
                # LoanAccountNo Validation (Null + Duplicate)
                # ================================
                if "loanaccountno" in col_map:
                    value = ws.cell(row=row, column=col_map["loanaccountno"]).value
                    val_str = str(value).strip() if value else ""
                    if val_str == "":
                        ws.cell(row=row, column=col_map["loanaccountno"]).fill = red_fill
                    else:
                        if val_str in unique_loan_acc:
                            ws.cell(row=row, column=col_map["loanaccountno"]).fill = red_fill
                        else:
                            unique_loan_acc.add(val_str)

                # ================================
                # AccountNo Validation (Null + Duplicate)
                # ================================
                if "accountno" in col_map:
                    value = ws.cell(row=row, column=col_map["accountno"]).value
                    val_str = str(value).strip() if value else ""
                    if val_str == "":
                        ws.cell(row=row, column=col_map["accountno"]).fill = red_fill
                    else:
                        if val_str in unique_accno:
                            ws.cell(row=row, column=col_map["accountno"]).fill = red_fill
                        else:
                            unique_accno.add(val_str)

                # ================================
                # MemberName Validation (Null + Duplicate)
                # ================================
                if "membername" in col_map:
                    value = ws.cell(row=row, column=col_map["membername"]).value
                    val_str = str(value).strip() if value else ""
                    if val_str == "":
                        ws.cell(row=row, column=col_map["membername"]).fill = red_fill
                    else:
                        if val_str in unique_membername:
                            ws.cell(row=row, column=col_map["membername"]).fill = red_fill
                        else:
                            unique_membername.add(val_str)

                # ================================
                # MemberID Validation (Null + Duplicate)
                # ================================
                if "memberid" in col_map:
                    value = ws.cell(row=row, column=col_map["memberid"]).value
                    val_str = str(value).strip() if value else ""
                    if val_str == "":
                        ws.cell(row=row, column=col_map["memberid"]).fill = red_fill
                    else:
                        if val_str in unique_memberid:
                            ws.cell(row=row, column=col_map["memberid"]).fill = red_fill
                        else:
                            unique_memberid.add(val_str)
        # Save file
        wb.save(file_path)
        print(f"Excel file updated successfully: {file_path}")
        os.startfile(file_path) # Open the file after saving(added)

    except PermissionError:
        print(f"Permission denied. Please close the file: {file_path}")
    except Exception as e:
        print(f"Error: {e}")

else:
    print("No file selected.")
