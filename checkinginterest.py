import os
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(
    title="Select Excel File", 
    filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
)
root.destroy()

if not file_path:
    print("No file selected. Exiting.")
    raise SystemExit

df = pd.read_excel(file_path, sheet_name="LoanMain", engine="openpyxl")

folder = os.path.dirname(file_path)
output_file = os.path.join(folder, "CorrectedFile.xlsx")

wb = load_workbook(file_path)
ws = wb["LoanMain"]

red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

column_to_check ="InterestRate"

if column_to_check not in df.columns:
    print(f"Column '{column_to_check}' not found in Excel.")
    raise SystemExit

col_idx = df.columns.get_loc(column_to_check) + 1  

for row_idx, value in enumerate(df[column_to_check], start=2): 
    if pd.isna(value) or not str(value).replace('.', '', 1).isdigit():
        ws.cell(row=row_idx, column=col_idx).fill = red

wb.save(output_file)
print(f"Done Highlighted invalid Interest Rate cells.\nSaved as: {output_file}")

try:
    os.startfile(output_file)
except Exception as e:
    print(f"Could not open file automatically: {e}")
