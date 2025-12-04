import os
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import re

root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(
    title="Select Excel File",
    filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
)
root.destroy()

if not file_path:
    print("No file selected. Exiting.")
    raise SystemExit

df = pd.read_excel(file_path, sheet_name="LoanMain", engine="openpyxl")

print(df.columns.tolist())

folder = os.path.dirname(file_path)
output_file = os.path.join(folder, "CorrectedFile.xlsx")

wb = load_workbook(file_path)
ws = wb["LoanMain"]

red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

date_columns = ["LoanIssueDate BS", "LoanIssueDate BS.1", "MaturityDateBS"]

date_pattern = re.compile(r"^\d{4}-(0[1-9]|1[0-2])-(0[1-9]|[12][0-9]|3[01])$")

for col in date_columns:

    if col not in df.columns:
        print(f"Column '{col}' not found in Excel. Skipping.")
        continue

    col_idx = df.columns.get_loc(col) + 1

    for row_idx, value in enumerate(df[col], start=2):
        str_value = str(value).strip() if not pd.isna(value) else ""

        if str_value == "":
            ws.cell(row=row_idx, column=col_idx).fill = red
            continue

        if str_value.replace('.', '', 1).isdigit():
            ws.cell(row=row_idx, column=col_idx).fill = red
            continue

        if not date_pattern.match(str_value):
            ws.cell(row=row_idx, column=col_idx).fill = red
            continue


wb.save(output_file)
print(f"\nDone highlighting invalid date cells.\nSaved as: {output_file}")

try:
    os.startfile(output_file)
except Exception as e:
    print(f"Could not open file automatically: {e}")
