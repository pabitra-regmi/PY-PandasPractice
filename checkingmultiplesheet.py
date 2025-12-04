import os
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(
    title="Select Excel File", 
    filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
)
root.destroy()

if not file_path:
    print("No file selected. Exiting.")
    raise SystemExit

wb = load_workbook(file_path)
sheet_names = wb.sheetnames

folder = os.path.dirname(file_path)
output_file = os.path.join(folder, "CorrectedFile_Output.xlsx")

red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

column_to_check = "InterestRate"

processed_sheets = []

for sheet in sheet_names:
    df = pd.read_excel(file_path, sheet_name=sheet, engine="openpyxl")
    df.columns = df.columns.str.strip()

    print(f"\nChecking sheet: {sheet}")
    print("Columns:", df.columns.tolist())

    if column_to_check not in df.columns:
        print(f"  Column '{column_to_check}' NOT found in this sheet.")
        continue  

    print(f" Column '{column_to_check}' found Processing..")

    ws = wb[sheet]
    col_idx = df.columns.get_loc(column_to_check) + 1  

    for row_idx, value in enumerate(df[column_to_check], start=2):

        value_str = str(value).strip()

        has_percent = "%" in value_str

        is_numeric = value_str.replace('.', '', 1).isdigit()

        if pd.isna(value) or not is_numeric or has_percent:
            ws.cell(row=row_idx, column=col_idx).fill = red

    processed_sheets.append(sheet)

wb.save(output_file)
print("\n")
print("Succesfully highlighted the invalid errors in InterestRate.")
print("Corrected file saved as:", output_file)
print("Sheets processed:", processed_sheets)

try:
    os.startfile(output_file)
except Exception as e:
    print(f"File can not be open: {e}")
