

#download pip install pyodbc

import os
import pandas as pd
import pyodbc
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

# Select Excel file
root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename(
    title="Select Client Excel File",
    filetypes=[("Excel files", "*.xlsx;*.xlsm")]
)
root.destroy()
if not file_path:
    raise SystemExit("No file selected")

# Connect to SQL Server
conn = pyodbc.connect(
    r'Driver={SQL Server};'
    r'Server=LAPTOP-CMR6SBT0\SQLEXPRESS;'
    r'Database=employee;' 
    r'Trusted_Connection=yes;'
)

# Read payment table
db_df = pd.read_sql("SELECT * FROM dbo.PaymentMethod", conn)

# Detect ID and text columns
id_column = None
text_columns = []
for col in db_df.columns:
    if pd.api.types.is_numeric_dtype(db_df[col]):
        id_column = col
    else:
        text_columns.append(col)
if id_column is None or not text_columns:
    raise SystemExit("Cannot detect ID or text column")

# Map aliases to the same DB ID
aliases = ["cash", "rakam", "paisa"]
# Find the DB ID of 'cash'
cash_id_row = db_df[text_columns].apply(lambda col: col.str.lower() == "cash").any(axis=1)
if not cash_id_row.any():
    raise SystemExit("Cannot find 'cash' in DB")
cash_id = db_df.loc[cash_id_row, id_column].iloc[0]
# Map all aliases to this ID
payment_lookup = {alias: cash_id for alias in aliases}

# Read Excel
file_df = pd.read_excel(file_path)
file_df["PaymentID"] = None

# Assign IDs based on aliases
for col in file_df.columns:
    mapped_ids = file_df[col].astype(str).str.strip().str.lower().map(payment_lookup)
    file_df["PaymentID"] = mapped_ids.combine_first(file_df["PaymentID"])

# Write IDs back to Excel
wb = load_workbook(file_path)
ws = wb.active
headers = [cell.value for cell in ws[1]]
if "PaymentID" in headers:
    id_col_idx = headers.index("PaymentID") + 1
else:
    id_col_idx = ws.max_column + 1
    ws.cell(row=1, column=id_col_idx).value = "PaymentID"

for row_idx, value in enumerate(file_df["PaymentID"], start=2):
    if pd.notna(value):
        ws.cell(row=row_idx, column=id_col_idx).value = int(value)

wb.save(file_path)
conn.close()

# Open updated Excel
os.startfile(file_path)
print("Done.")
