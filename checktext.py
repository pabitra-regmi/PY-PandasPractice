import os
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(
    title="Select Excel File", 
    filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
)
root.destroy()

if not file_path:
    print("No file selected. Exiting.")
    raise SystemExit

df = pd.read_excel(file_path, engine="openpyxl")
sheet_name = df.columns.name if df.columns.name else df.columns[0]

folder = os.path.dirname(file_path)
output_file = os.path.join(folder, "CorrectedFile.xlsx")

wb = load_workbook(file_path)
ws = wb[wb.sheetnames[0]] 

red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for col_idx, col_name in enumerate(df.columns, start=1):
   
    numeric_ratio = pd.to_numeric(df[col_name], errors='coerce').notna().mean()
    if numeric_ratio < 0.5:
        for row_idx, value in enumerate(df[col_name], start=2):  
            if pd.notna(value) and re.search(r'\d', str(value)):
                ws.cell(row=row_idx, column=col_idx).fill = red

wb.save(output_file)
print(f"Done Highlighted text cells containing numbers.\nSaved as: {output_file}")

try:
    os.startfile(output_file)
except Exception as e:
    print(f"Could not open file automatically: {e}")
