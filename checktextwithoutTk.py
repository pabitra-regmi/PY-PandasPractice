import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

file_path = input("Enter full path of your Excel file: ").strip()

if not file_path or not os.path.exists(file_path):
    print("File not found. Exiting.")
    raise SystemExit

df = pd.read_excel(file_path, engine="openpyxl")

folder = os.path.dirname(file_path)
output_file = os.path.join(folder, "TextColumnsWithNumbers_Highlighted.xlsx")

wb = load_workbook(file_path)
ws = wb[wb.sheetnames[0]]  

red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for col_idx, col_name in enumerate(df.columns, start=1):

    numeric_ratio = pd.to_numeric(df[col_name], errors='coerce').notna().mean()
    if numeric_ratio < 0.5:
        for row_idx, value in enumerate(df[col_name], start=2):  
            if pd.notna(value) and re.search(r'\d', str(value)):
                ws.cell(row=row_idx, column=col_idx).fill = red

wb.save(output_file)
print(f"Done Highlighted text cells containing numbers.\nSaved as: {output_file}")

try:
    os.startfile(output_file)
except Exception as e:
    print(f"Could not open file automatically: {e}")
